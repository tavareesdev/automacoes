from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import time
import os
import glob
import re
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl import Workbook

edge_driver_path = 'caminho_para_o_executavel_do_edgedriver.exe'

edge_options = webdriver.EdgeOptions()
edge_options.use_chromium = True

navegador = webdriver.Edge(options=edge_options)

navegador.get("https://administracao.evoluaprofissional.com.br/")

campo_email = navegador.find_elements(By.XPATH, '//*[@id="Login"]')
campo_email[0].send_keys("easytech.osasco@gmail.com")
time.sleep(2)

campo_senha = navegador.find_elements(By.XPATH, '//*[@id="Senha"]')
campo_senha[0].send_keys("$ucesso3331")
campo_senha[0].send_keys(Keys.ENTER)
time.sleep(2)

iframe = WebDriverWait(navegador, 10).until(
    EC.presence_of_element_located((By.ID, "huggy-trigger-27202"))
    and EC.visibility_of_element_located((By.ID, "huggy-trigger-27202"))
)

navegador.switch_to.frame(iframe)

navegador.find_element(By.XPATH, '/html/body/div[2]/div/div/header/div[1]/div').click()

navegador.switch_to.default_content()
time.sleep(2)

navegador.find_element(By.XPATH, '//*[@id="accordionMenu"]/a[6]').click()
time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="pedagogico"]/ul/li[11]/a').click()
time.sleep(2)

data1 = navegador.find_elements(By.ID, "dt_inicial")
data_de_hoje = datetime.now().date()

if datetime.now().weekday() == 0:
    dia_semana = 'SÁBADO'
elif datetime.now().weekday() == 1:
    dia_semana = 'SEGUNDA-FEIRA'
elif datetime.now().weekday() == 2:
    dia_semana = 'TERÇA-FEIRA'
elif datetime.now().weekday() == 3:
    dia_semana = 'QUARTA-FEIRA'
elif datetime.now().weekday() == 4:
    dia_semana = 'QUINTA-FEIRA'
elif datetime.now().weekday() == 5:
    dia_semana = 'SEXTA-FEIRA'
else:
    dia_semana = 'DOMINGO'

if dia_semana == 'SÁBADO':
    data_de_ontem = data_de_hoje - timedelta(days=2)
else:
    data_de_ontem = data_de_hoje - timedelta(days=1)

data_de_ontem = data_de_ontem.strftime("%d/%m/%Y")
data1[0].send_keys(data_de_ontem)

data2 = navegador.find_elements(By.ID, "dt_final")
data2[0].send_keys(data_de_ontem)
navegador.find_element(By.XPATH, '/html').send_keys(Keys.ESCAPE)

navegador.find_element(By.XPATH, '//*[@id="data"]/div[2]/div/button').click()

# Página Situação
time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="situacao"]/div/div/div[1]/div/div/div/div[2]/a').click()

time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="situacao"]/div/div/div[2]/div/button').click()

# Página Organograma
time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="organograma"]/button').click()

time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="organograma"]/div[2]/div/button').click()

# Página Alunos
time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="alunos"]/button').click()

time.sleep(2)
navegador.find_element(By.XPATH, '//*[@id="alunos"]/div[2]/div[1]/button').click()

# Aguarda o download ser concluído
time.sleep(10)

# Obtém o diretório de download
diretorio_download = "C:\\Users\\User\\Downloads"

arquivos = os.listdir(diretorio_download)

# Seleciona o arquivo mais recente com extensão .xlsx
arquivo_excel = max([f for f in arquivos if f.endswith('.xlsx')], key=lambda x: os.path.getmtime(os.path.join(diretorio_download, x)))

# Caminho completo do arquivo Excel
caminho_arquivo_excel = os.path.join(diretorio_download, arquivo_excel)

# Lê o arquivo Excel usando o pandas
df = pd.read_excel(caminho_arquivo_excel)

def trim_values(value):
    if isinstance(value, str):
        # Substituir dois espaços juntos por um espaço
        return re.sub('  +', ' ', value.strip())
    return value

# Ler os nomes da coluna D e armazenar em variáveis
variables = {}
for i, value in enumerate(df['Aluno']):
    if value not in variables.values():
        variables[f'var{i+1}'] = value

# Carregar o arquivo do Excel
excel_file = 'C:\\Users\\User\\Documents\\HORARIOS OSASCO NOVO.xlsx'

# Ler a planilha do Excel
df_excel = pd.read_excel(excel_file, sheet_name=dia_semana)

# Lista para armazenar os resultados da comparação
resultados = {}

# Colunas a serem comparadas (B até I)
colunas_comparar = df_excel.columns[1:9] 

# Ler os nomes das colunas B até I e armazenar em variáveis
columns = df_excel.columns[1:9]  # Colunas B até I (0-indexed)
variables2 = {}
for i, row in df_excel.iterrows():
    if i < 5:  # Pular as primeiras 5 linhas
        continue
    for j, value in enumerate(row[1:9], start=1):
        var_name = f'var{i+1}_{j}'
        if value not in variables2.values():
            variables2[var_name] = value

# Carregar o arquivo do Excel
excel_file2 = 'C:\\Users\\User\\Documents\\_LIGAÇÕES FALTANTES NOVO 2024.xlsx'

# Ler a planilha do Excel
df_excel2 = pd.read_excel(excel_file2, sheet_name='ABRIL')

# Lista para armazenar os resultados da comparação
resultados2 = {}

# Selecionar os nomes da coluna A onde o valor da coluna E é "NÃO"
alunos_faltantes = df_excel2[(df_excel2.iloc[:, 4] == 'NÃO') & (df_excel2.iloc[:, 1] == 'ATIVO')].iloc[:, 0].drop_duplicates().tolist()

# Criar a estrutura de dados desejada
variables3 = {f'var{i+1}': aluno for i, aluno in enumerate(alunos_faltantes)}

def clean_text(text):
    if isinstance(text, str):
        mapping = {
            'Á': 'A', 'À': 'A', 'Ã': 'A', 'Â': 'A', 'Ä': 'A',
            'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
            'Í': 'I', 'Ì': 'I', 'Î': 'I', 'Ï': 'I',
            'Ó': 'O', 'Ò': 'O', 'Õ': 'O', 'Ô': 'O', 'Ö': 'O',
            'Ú': 'U', 'Ù': 'U', 'Û': 'U', 'Ü': 'U',
            'Ç': 'C',
            'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a', 'ä': 'a',
            'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
            'í': 'i', 'ì': 'i', 'î': 'i', 'ï': 'i',
            'ó': 'o', 'ò': 'o', 'õ': 'o', 'ô': 'o', 'ö': 'o',
            'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
            'ç': 'c'
        }
        cleaned_text = ''.join(mapping.get(char, char) for char in text)
        return cleaned_text.upper()
    else:
        return text


# Aplicar a função trim_values em todos os valores da array
variables = {key: trim_values(value) for key, value in variables.items()}
variables2 = {key: trim_values(value) for key, value in variables2.items()}
variables3 = {key: trim_values(value) for key, value in variables3.items()}

variables = {key: clean_text(value) for key, value in variables.items()}
variables2 = {key: clean_text(value) for key, value in variables2.items()}
variables3 = {key: clean_text(value) for key, value in variables3.items()}

# Convertendo os valores do dicionário em uma lista
variables = list(variables.values())
variables2 = list(variables2.values())
variables3 = list(variables3.values())

variables.sort()
variables2.sort(key=str)
variables3.sort(key=str)

# Encontrar os valores que estão em ambas as listas
Presentes = list(set(variables).intersection(variables2))
Presentes.sort(key=str)

set1 = set(variables)
set2 = set(variables2)
set3 = set(variables3)

# Encontrar os valores que estão em ambas as listas
Reposicoes = list(set(variables).intersection(variables3))
Reposicoes.sort(key=str)

valores_diferentes = list(set1.symmetric_difference(set2))

Todos_presentes1 = Presentes + Reposicoes

Todos_presentes = []
for item in Todos_presentes1:
    if item not in Todos_presentes:
        Todos_presentes.append(item)

set4 = set(Todos_presentes)
set5 = set(Reposicoes)

Faltantes = []

for item in variables2:
    if item not in Todos_presentes:
        Faltantes.append(item)

Faltantes.sort(key=str)

a = True

os.remove(caminho_arquivo_excel)


data_faltantes = {}

# for item in Faltantes:

#     navegador.find_element(By.XPATH, '//*[@id="pedagogico"]/ul/li[11]/a').click()
#     time.sleep(2)

#     data1 = navegador.find_elements(By.ID, "dt_inicial")
#     data_antiga = '17/10/2005'
#     data1[0].send_keys(data_antiga)

#     data2 = navegador.find_elements(By.ID, "dt_final")
#     data2[0].send_keys(data_de_ontem)
#     navegador.find_element(By.XPATH, '/html').send_keys(Keys.ESCAPE)
#     time.sleep(3)

#     navegador.find_element(By.XPATH, '//*[@id="data"]/div[2]/div/button').click()

#     # Página Situação
#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="situacao"]/div/div/div[1]/div/div/div/div[2]/a').click()

#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="situacao"]/div/div/div[2]/div/button').click()

#     # Página Organograma
#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="organograma"]/button').click()

#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="organograma"]/div[2]/div/button').click()

#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="heading-0"]/div').click()

#     campo_alunos = navegador.find_elements(By.XPATH, '//*[@id="datatable-0_filter"]/label/input')
#     campo_alunos[0].send_keys(item)
#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="alunos"]/button').click()
#     time.sleep(2)
#     navegador.find_element(By.XPATH, '//*[@id="alunos"]/div[2]/div[1]/button').click()
#     # Obtém o diretório de download
#     diretorio_download = "C:\\Users\\User\\Downloads"

#     arquivos = os.listdir(diretorio_download)

#     # Seleciona o arquivo mais recente com extensão .xlsx
#     arquivo_excel = max([f for f in arquivos if f.endswith('.xlsx')], key=lambda x: os.path.getmtime(os.path.join(diretorio_download, x)))

#     # Caminho completo do arquivo Excel
#     caminho_arquivo_excel = os.path.join(diretorio_download, arquivo_excel)

#     # Lê o arquivo Excel usando o pandas
#     df_aluno = pd.read_excel(caminho_arquivo_excel)
#     print(df_aluno)
    
#     ultimo_valor = df_aluno['Data'].iloc[-1]

#     data_faltantes[f'var{i+1}'] = ultimo_valor + ' ' + item

#     os.remove(caminho_arquivo_excel)

# Criar um novo arquivo de Excel
wb = load_workbook('C:\\Users\\User\\Documents\\Relatórios\\Relatório Abril.xlsx')

# Data de ontem
data_de_ontem = (datetime.now() - timedelta(days=1)).strftime('%d-%m-%Y')

# Criar uma nova planilha
nova_planilha = wb.create_sheet(data_de_ontem)

# Adicionar cabeçalho
nova_planilha['A1'] = 'Presentes'
nova_planilha['B1'] = 'Reposições'
nova_planilha['C1'] = 'Faltantes'

# Adicionar dados da lista
for idx, nome in enumerate(Presentes, start=2):
    nova_planilha[f'A{idx}'] = nome

# Adicionar dados da lista
for idx, nome in enumerate(Reposicoes, start=2):
    nova_planilha[f'B{idx}'] = nome

# Adicionar dados da lista
for idx, nome in enumerate(Faltantes, start=2):
    nova_planilha[f'C{idx}'] = nome

# Criar o diretório se ele não existir
if not os.path.exists('C:\\Users\\User\\Documents\\Relatórios'):
    os.makedirs('C:\\Users\\User\\Documents\\Relatórios')

wb.save(f'C:\\Users\\User\\Documents\\Relatórios\\Relatório Abril.xlsx')

navegador.quit()